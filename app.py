from flask import Flask,jsonify,request
from flask_sqlalchemy import SQLAlchemy
from flask_jwt_extended import create_access_token,JWTManager,get_jwt_identity,jwt_required
from flask_restful import Api,Resource
from openpyxl import load_workbook
from io import BytesIO
from datetime import timedelta

app=Flask(__name__)
app.config['SECRET_KEY']='thissecretkey'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///site.db'
api=Api(app)
db=SQLAlchemy(app)
jwt=JWTManager(app)


class User(db.Model):
    id=db.Column(db.Integer(),primary_key=True)
    username=db.Column(db.String(200))
    password=db.Column(db.String(200))
    
class UserRegistration(Resource):
    def post(self):
        data=request.get_json()
        username=data['username']
        password=data['password']

        if not username and not password:
            return jsonify({'message':'username and password are invalid'}),400
        if User.query.filter_by(username=username).first():
            return jsonify({'message':'username allredy exsit'})
        new_user=User(username=username,password=password)
        db.session.add(new_user)
        db.session.commit()
        return jsonify({'message':'new user created sucsessfully'})

class UserLogin(Resource):
    def post(self):
        data = request.get_json()
        username = data['username']
        password = data['password']

        user = User.query.filter_by(username=username).first()
        print(user)
        if user and user.password==password :
            access_token = create_access_token(identity=user.id,expires_delta=timedelta(hours=1))
            print(access_token)
            return {'access_token': access_token}, 200

        return {'message': 'Invalid credentials'}, 401

class UserProtected(Resource):
    @jwt_required()
    def get(self):
        curent_user=get_jwt_identity()
        return jsonify({'message':f'hello {curent_user}your resourse are protected'}),200




class ParcelExcle(Resource):
    def post(self):
        try:
            data = request.files['file']
            workbook = load_workbook(data)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    username, password = row
                    data2 = User(username=username, password=password)
                    db.session.add(data2)
                # data2.save()
                except ValueError as e:
                   print(f"Error unpacking row: {e}")
            db.session.commit()
            return jsonify({'message': 'ok'}), 200
        except Exception as e:
            return jsonify({'error': str(e)}), 500



class GetUser(Resource):
    @jwt_required()
    def get(self,user_id):
        try:
            user=User.query.get(user_id)
            if user:
                print(user)
                return jsonify({'username':user.username,'password':user.password}),200
                
            else:
                return jsonify({'error':'user not found'}),400
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error':str(e)}),500  


class UpdateUser(Resource):
    @jwt_required()
    def put(self,user_id):
        try:
            user=User.query.get(user_id)
            if user:
                data=request.get_json()
                username=data.get('username')
                password=data.get('password')
                if username:
                    user.username=username
                if password:
                    user.password=password
                db.session.commit()
                return jsonify({'message':'user is update sucssefully'}),200
            else:
                return jsonify({'error':'user not found'}),400
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error':str(e)}),500
            
class DeleteUser(Resource):
    @jwt_required()
    def delete(self,user_id):
        try:
            user = User.query.get(user_id)
            print(f"User: {user}") 
            if user:
                db.session.delete(user)
                db.session.commit()
                return jsonify({'message': 'User deleted successfully'}), 200
            else:
                return jsonify({'error': 'User not found'}), 404
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': str(e)}), 500





api.add_resource(UserRegistration,'/register')
api.add_resource(UserLogin,'/login')
api.add_resource(UserProtected,'/protec')
api.add_resource(ParcelExcle,'/parce')
api.add_resource(GetUser,'/getuser/<int:user_id>')
api.add_resource(UpdateUser,'/updatuser/<int:user_id>')
api.add_resource(DeleteUser, '/delete_user/<int:user_id>')


if __name__=='__main__':
    with app.app_context():
        db.create_all()
        app.run(debug=True)
